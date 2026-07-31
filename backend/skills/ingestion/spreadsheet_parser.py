"""XLSX/CSV ingestion — each sheet (or CSV) becomes one RawPage."""

from __future__ import annotations

import csv
from pathlib import Path

from backend.skills.ingestion.pdf_parser import RawPage, RawTextBlock
from backend.skills.ingestion.text_parser import TextDocument


def parse_spreadsheet(path: str | Path) -> TextDocument:
    path = Path(path)
    ext = path.suffix.lower()
    if ext == ".csv":
        return _parse_csv(path)
    if ext in {".xlsx", ".xlsm"}:
        return _parse_xlsx(path)
    raise ValueError(f"Unsupported spreadsheet: {ext}")


def _parse_csv(path: Path) -> TextDocument:
    with path.open(encoding="utf-8", errors="replace", newline="") as f:
        rows = list(csv.reader(f))
    return TextDocument(pages=[_sheet_to_page(1, path.stem, rows)])


def _parse_xlsx(path: Path) -> TextDocument:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required for XLSX ingestion. pip install openpyxl"
        ) from e

    wb = load_workbook(str(path), data_only=True, read_only=True)
    pages: list[RawPage] = []
    for i, name in enumerate(wb.sheetnames, start=1):
        ws = wb[name]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c).strip() for c in row]
            if any(cells):
                rows.append(cells)
        pages.append(_sheet_to_page(i, name, rows))
    wb.close()
    if not pages:
        pages.append(_sheet_to_page(1, "empty", []))
    return TextDocument(pages=pages)


def _sheet_to_page(page_number: int, title: str, rows: list[list[str]]) -> RawPage:
    width, height = 612.0, 792.0
    lines = [f"# {title}"] if title else []
    table_lines = [" | ".join(r) for r in rows]
    lines.extend(table_lines)
    full = "\n".join(lines).strip()

    blocks: list[RawTextBlock] = []
    if title:
        blocks.append(
            RawTextBlock(
                text=title,
                bbox=(36.0, 36.0, width - 36.0, 60.0),
                block_type="TEXT",
                confidence=1.0,
            )
        )
    if table_lines:
        # Cap very large sheets for storage; keep head+tail signal
        max_rows = 200
        if len(table_lines) > max_rows:
            shown = (
                table_lines[:150]
                + [f"... ({len(table_lines) - 180} rows omitted) ..."]
                + table_lines[-30:]
            )
        else:
            shown = table_lines
        blocks.append(
            RawTextBlock(
                text="\n".join(shown),
                bbox=(36.0, 70.0, width - 36.0, height - 36.0),
                block_type="TABLE",
                confidence=1.0,
            )
        )

    if len(rows) > 200:
        full_lines = [f"# {title}"] + table_lines[:150] + table_lines[-30:]
        full = "\n".join(full_lines)

    return RawPage(
        page_number=page_number,
        width=width,
        height=height,
        text_layer_available=bool(full),
        full_text=full,
        blocks=blocks,
        image_count=0,
        drawing_count=0,
    )
